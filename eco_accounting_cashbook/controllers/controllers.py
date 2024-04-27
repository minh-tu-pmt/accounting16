# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request
import json

import openpyxl
from io import BytesIO
from openpyxl.styles import Font,Alignment,Border, Side
import base64
import os
from datetime import datetime


class ActAccountingCashbook(http.Controller):

    @http.route('/eco_accounting_cashbook/objects/account_company', auth='public', type="json")
    def account_company(self,account_id, **kwargs):
        account_id = request.env['account.account'].sudo().browse(account_id)
        return {
            'company': account_id.company_id.name,
            'code': account_id.code
        }   

    @http.route('/eco_accounting_cashbook/objects/action_report', auth='public', type="json")
    def action_report(self,account_id,from_date, to_date, **kwargs):
        # account_id = 196
        rs = request.env['account.account'].sudo().search([('id', 'child_of', account_id)])
        # rs = rs.filtered(lambda acc: acc.id != account_id)
        
        parent_ids = rs.mapped('parent_id')
        rs = rs.filtered(lambda acc: acc.id not in parent_ids.ids)

        move_line_ids = request.env['account.move.line'].sudo().search([('account_id', 'in', rs.ids), ('date', '<', from_date), ('parent_state', '=','posted')])
        debit = sum(move_line_ids.mapped('debit'))
        credit = sum(move_line_ids.mapped('credit'))

        remain = debit - credit
        move_ids = request.env['account.move.line'].sudo().search([('account_id', 'in', rs.ids), ('date', '>=', from_date), ('date', '<=', to_date), ('parent_state', '=','posted')])

        result = []
        for m in move_ids:
            result.append({
                'date': m.date.strftime('%d/%m/%Y'),
                'ref': m.ref or '',
                'move_name': m.move_name,
                'debit': m.debit,
                'credit': m.credit,
                'company': m.company_id.name,
                'move_id': m.move_id.id,
            })
        return {'out_debit': remain, 'move_lines': result}
    
    @http.route('/eco_accounting_cashbook/objects/export_excel', auth='public', type="json")
    def export_excel(self, account_id,from_date, to_date, **kwargs):
        path=os.path.join(os.path.dirname(__file__), '../data','report_cashbook.xlsx')
        wb = openpyxl.load_workbook(path)
        worksheet = wb.active

        center_aligned_text = Alignment(horizontal='center', vertical='center')
        font_time_12 = Font(name='Times New Roman', size=12)
        font_time_11 = Font(name='Times New Roman', size=11)
        bold_font_12 = Font(name='Times New Roman', bold=True, size=12)
        bold_font_11 = Font(name='Times New Roman', bold=True, size=11)
        # Tạo đối tượng Side để định nghĩa kiểu viền
        thin_border = Side(border_style="medium", color="000000")

        # Tạo đối tượng Border và áp dụng đối tượng Side đã tạo
        border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

        rs = self.action_report(account_id,from_date, to_date)
        acc = request.env['account.account'].sudo().browse(account_id)
        worksheet['A5'] = 'Tài khoản: %s'%(acc.code)
        worksheet['A1'] = '%s'%(acc.company_id.name)
        worksheet['A6'] = 'Từ ngày %s đến ngày %s'%(datetime.strptime(from_date, '%Y-%m-%d').strftime('%d-%m-%Y'), datetime.strptime(to_date, '%Y-%m-%d').strftime('%d-%m-%Y'))

        out_debit = rs.get('out_debit')
        worksheet['I7'] = out_debit


        move_lines = rs.get('move_lines')

        line = 10
        money = out_debit
        debit = 0
        credit = 0
        for move in move_lines:
            worksheet['A%s'%line] = move.get('company')
            worksheet['A%s'%line].font = font_time_12
            worksheet['A%s'%line].border = border

            worksheet['B%s'%line] = move.get('date')
            worksheet['B%s'%line].font = font_time_12
            worksheet['B%s'%line].border = border

            worksheet['C%s'%line] = move.get('date')
            worksheet['C%s'%line].font = font_time_12
            worksheet['C%s'%line].border = border

            if move.get('debit'):
                worksheet['D%s'%line] = move.get('move_name')
                worksheet['D%s'%line].font = font_time_12
                worksheet['D%s'%line].border = border
                worksheet['E%s'%line].border = border
            else:
                worksheet['E%s'%line] = move.get('move_name')
                worksheet['E%s'%line].font = font_time_12
                worksheet['E%s'%line].border = border
                worksheet['D%s'%line].border = border

            worksheet['F%s'%line] = move.get('ref')
            worksheet['F%s'%line].font = font_time_12
            worksheet['F%s'%line].border = border

            if move.get('debit'):
                worksheet['G%s'%line] = move.get('debit')
                worksheet['G%s'%line].font = font_time_12
                worksheet['G%s'%line].border = border
                worksheet['H%s'%line].border = border
            else:
                worksheet['H%s'%line] = move.get('credit')
                worksheet['H%s'%line].font = font_time_12
                worksheet['H%s'%line].border = border
                worksheet['G%s'%line].border = border

            # worksheet['I%s'%line] = money+move.get('debit')-move.get('credit')
            money = money+move.get('debit')-move.get('credit')
            worksheet['I%s'%line] = money
            worksheet['I%s'%line].font = font_time_12
            worksheet['I%s'%line].border = border
            worksheet['J%s'%line].border = border

            debit = debit + move.get('debit')
            credit = credit + move.get('credit')
            worksheet.insert_rows(idx=line + 1, amount=1)
            line = line+1

        worksheet['H%s'%line] = 'TỔNG PHÁT SINH NỢ'
        worksheet['H%s'%line].font = bold_font_12
        worksheet['H%s'%line].alignment = center_aligned_text

        worksheet['I%s'%line] = debit
        worksheet['I%s'%line].font = bold_font_12

        line = line+1
        worksheet['H%s'%line] = 'TỔNG PHÁT SINH CÓ'
        worksheet['H%s'%line].font = bold_font_12
        worksheet['H%s'%line].alignment = center_aligned_text

        worksheet['I%s'%line] = credit
        worksheet['I%s'%line].font = bold_font_12

        line=line+1
        worksheet['H%s'%line] = 'SỐ TỒN CUỐI'
        worksheet['H%s'%line].font = bold_font_12
        worksheet['H%s'%line].alignment = center_aligned_text

        worksheet['I%s'%line] = money
        worksheet['I%s'%line].font = bold_font_12
        line = line + 1
        
        worksheet.merge_cells('H%s:I%s'%(line, line))
        worksheet['H%s'%line] = 'Ngày .... tháng .... năm ....'
        worksheet['H%s'%line].font = bold_font_11
        worksheet['H%s'%line].alignment = center_aligned_text

        line = line + 1
        worksheet.merge_cells('A%s:B%s'%(line, line))
        worksheet['A%s'%line] = 'Người ghi sổ'
        worksheet['A%s'%line].font = bold_font_11
        worksheet['A%s'%line].alignment = center_aligned_text

        worksheet.merge_cells('D%s:E%s'%(line, line))
        worksheet['D%s'%line] = 'Kế toán trưởng'
        worksheet['D%s'%line].font = bold_font_11
        worksheet['D%s'%line].alignment = center_aligned_text

        worksheet.merge_cells('H%s:I%s'%(line, line))
        worksheet['H%s'%line] = 'Giám đốc'
        worksheet['H%s'%line].font = bold_font_11
        worksheet['H%s'%line].alignment = center_aligned_text

        line = line + 1
        worksheet.merge_cells('A%s:B%s'%(line, line))
        worksheet['A%s'%line] = '(Ký, họ tên)'
        worksheet['A%s'%line].font = font_time_11
        worksheet['A%s'%line].alignment = center_aligned_text

        worksheet.merge_cells('D%s:E%s'%(line, line))
        worksheet['D%s'%line] = '(Ký, họ tên)'
        worksheet['D%s'%line].font = font_time_11
        worksheet['D%s'%line].alignment = center_aligned_text

        worksheet.merge_cells('H%s:I%s'%(line, line))
        worksheet['H%s'%line] = '(Ký, họ tên, đóng dấu)'
        worksheet['H%s'%line].font = font_time_11
        worksheet['H%s'%line].alignment = center_aligned_text

        output = BytesIO()

        # # Lưu workbook vào đối tượng BytesIO
        wb.save(output)

        # # Đặt con trỏ ở vị trí đầu tiên
        output.seek(0)
        
        return {
            'type': 'ir.actions.act_url',
            'url': 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + base64.b64encode(output.read()).decode('utf-8'),
            'target': 'new',
            'file_name': 'report_cashbook'
        }
